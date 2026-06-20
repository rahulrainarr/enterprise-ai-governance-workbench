from __future__ import annotations

import csv
import json
from dataclasses import dataclass
from pathlib import Path
from typing import Any

from app.core.security import ensure_ingestion_allowed


@dataclass(frozen=True)
class AIMMAssessmentRecord:
    organization_name: str
    industry: str = ""
    assessment_date: str = ""
    data_readiness_score: float = 0.0
    infrastructure_readiness_score: float = 0.0
    cybersecurity_readiness_score: float = 0.0
    governance_maturity_score: float = 0.0
    ai_use_case_maturity_score: float = 0.0
    blockers: str = ""
    recommendations: str = ""
    priority_actions: str = ""


def import_aimm_records(path: str | Path) -> list[AIMMAssessmentRecord]:
    source = ensure_ingestion_allowed(path)
    suffix = source.suffix.lower()
    if suffix == ".json":
        raw = json.loads(source.read_text(encoding="utf-8"))
        rows = raw if isinstance(raw, list) else [raw]
        return [_record_from_mapping(row) for row in rows]
    if suffix == ".csv":
        with source.open(encoding="utf-8-sig", newline="") as handle:
            return [_record_from_mapping(row) for row in csv.DictReader(handle)]
    if suffix == ".xlsx":
        try:
            import openpyxl
        except Exception as exc:
            raise RuntimeError("openpyxl is required for AIMM XLSX imports.") from exc
        workbook = openpyxl.load_workbook(source, read_only=True, data_only=True)
        sheet = workbook.active
        headers = [str(cell.value or "").strip() for cell in next(sheet.iter_rows(max_row=1))]
        rows = [{header: value for header, value in zip(headers, row)} for row in sheet.iter_rows(min_row=2, values_only=True)]
        return [_record_from_mapping(row) for row in rows]
    raise ValueError(f"Unsupported AIMM import type: {suffix}")


def _record_from_mapping(row: dict[str, Any]) -> AIMMAssessmentRecord:
    normalized = {str(key).strip().lower().replace(" ", "_").replace("-", "_"): value for key, value in row.items()}
    return AIMMAssessmentRecord(
        organization_name=str(normalized.get("organization_name") or normalized.get("organization") or ""),
        industry=str(normalized.get("industry") or ""),
        assessment_date=str(normalized.get("assessment_date") or ""),
        data_readiness_score=_num(normalized.get("data_readiness_score")),
        infrastructure_readiness_score=_num(normalized.get("infrastructure_readiness_score")),
        cybersecurity_readiness_score=_num(normalized.get("cybersecurity_readiness_score")),
        governance_maturity_score=_num(normalized.get("governance_maturity_score")),
        ai_use_case_maturity_score=_num(normalized.get("ai_use_case_maturity_score")),
        blockers=str(normalized.get("blockers") or ""),
        recommendations=str(normalized.get("recommendations") or ""),
        priority_actions=str(normalized.get("priority_actions") or ""),
    )


def _num(value: Any) -> float:
    try:
        return float(value or 0)
    except (TypeError, ValueError):
        return 0.0
